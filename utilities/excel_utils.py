import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row
def get_col_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column
def read_data(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num,col_num).value
def write_data(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num,col_num).value = data
    workbook.save(file)
def fill_green(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    greenfill = PatternFill(start_color='00FF00', end_color='00FF00', patternType='solid')
    sheet.cell(row_num,col_num).fill = greenfill
    workbook.save(file)
def fill_red(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    redfill = PatternFill(start_color='FF0000', end_color='FF0000', patternType='solid')
    sheet.cell(row_num,col_num).fill = redfill
    workbook.save(file)